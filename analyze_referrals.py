import os
import json
import glob
import re
import pandas as pd
import requests
import PyPDF2
from urllib.parse import urlparse, parse_qs
from openai import AzureOpenAI
from datetime import datetime
import shutil
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Constants
RESUMES_DIR = "resumes"
EXCEL_OUTPUT = "referral_requests.xlsx"

# Azure OpenAI Configuration
client = AzureOpenAI(
    api_key="1354f5fafe564dd9952a16f42b02fa8a",
    azure_endpoint="https://agent42test.openai.azure.com/",
    api_version="2024-02-15-preview"
)

def get_latest_file(prefix):
    """Get the most recent file with the given prefix."""
    files = glob.glob(f"{prefix}_*.json")
    if not files:
        return None
    
    # Sort by creation time (newest first)
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def analyze_message_with_gpt(message_content):
    """
    Analyze a message using GPT-4 to determine if it's a job referral request
    and extract key information.
    """
    prompt = f"""
    Analyze the following LinkedIn message and determine:
    1. Is this a job referral request? (yes/no)
    2. If yes, extract the following information:
       - Sender's full name
       - Job position they're applying for
       - Job ID (if mentioned)
       - Company name (if mentioned)
       - Google Drive or other resume links (if any)
       - Email address (if any)
    
    Return the information in a structured JSON format with these fields:
    {{
        "is_referral_request": true/false,
        "sender_name": "Full Name",
        "job_position": "Position",
        "job_id": "ID123456",
        "company": "Company Name",
        "resume_links": ["link1", "link2"],
        "email": "email@example.com"
    }}

    If it's not a job referral request, just return {{"is_referral_request": false}}.
    
    LinkedIn Message:
    {message_content}
    """
    
    return call_gpt_api(prompt, "You are an assistant that analyzes LinkedIn messages for job referral requests and extracts structured information. Return only valid JSON.")

def check_resume_with_gpt(message_content):
    """
    Check if a message contains a resume attachment or link using GPT.
    """
    prompt = f"""
    Analyze the following LinkedIn message and determine:
    1. Does this message have a resume attached or linked? (yes/no)
    2. If yes, identify the type:
       - Google Drive link to resume
       - Direct PDF attachment mention
       - Other resume link type
    
    Return the information in a structured JSON format with these fields:
    {{
        "has_resume": true/false,
        "resume_type": "google_drive"/"pdf_attachment"/"other_link"/"none",
        "resume_link": "extracted link if available"
    }}
    
    LinkedIn Message:
    {message_content}
    """
    
    return call_gpt_api(prompt, "You are an assistant that analyzes whether LinkedIn messages have resume attachments or links. Return only valid JSON.", default_key="has_resume")

def extract_info_from_resume_with_gpt(text, filename=""):
    """Extract only the required info from resume text with GPT."""
    print(f"Extracting info from resume: {filename}")
    prompt = f"""
    Extract ONLY the following information from this resume:
    1. Full name
    2. Email address
    3. Phone number
    4. Total years of work experience (calculate this based on work history)
    
    For calculating years of experience:
    - Sum the total time (in years) across all professional work experiences
    - For current positions, calculate time from start date until present
    - For past positions, calculate time from start date to end date
    - Round to the nearest quarter (0.25, 0.5, 0.75, 1.0)
    - Do NOT include internships, volunteering, or educational experiences unless they were full-time paid professional roles
    - For overlapping time periods, do not double-count the overlapping time
    - If exact dates aren't provided, make a reasonable estimation based on the information available
    - Be consistent: if you calculate 1 year and 2 months, that should be 1.25 years
    
    Return the information in a structured JSON format with these fields:
    {{
        "name": "Full Name",
        "email": "email@example.com",
        "phone": "+1234567890",
        "years_of_experience": 5.25
    }}
    
    Resume:
    {text}
    """
    
    result = call_gpt_api(prompt, "You are an assistant that extracts specific information from resumes. Return only valid JSON.", default_key="success")
    
    # If we got a name back, log it
    if "name" in result and result["name"]:
        print(f"✅ Extracted name from resume: {result['name']}")
    
    return result

def call_gpt_api(prompt, system_content, default_key="is_referral_request"):
    """Generic function to call GPT API and handle responses."""
    try:
        response = client.chat.completions.create(
            model="gpt4o",
            messages=[
                {"role": "system", "content": system_content},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1  # Lower temperature for more consistent results
        )
        
        result = response.choices[0].message.content.strip()
        
        # Extract only the JSON part if there's additional text
        json_match = re.search(r'({.*})', result, re.DOTALL)
        if json_match:
            result = json_match.group(1)
            
        # Parse the JSON result
        try:
            return json.loads(result)
        except json.JSONDecodeError:
            print(f"Error parsing JSON response: {result}")
            return {default_key: False, "error": "Failed to parse response"}
            
    except Exception as e:
        print(f"Error calling GPT API: {str(e)}")
        return {default_key: False, "error": str(e)}

def extract_text_from_pdf(pdf_file):
    """Extract text content from a PDF file."""
    try:
        text = ""
        with open(pdf_file, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            num_pages = len(reader.pages)
            
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text += page.extract_text()
        
        return text
    except Exception as e:
        print(f"Error extracting text from PDF {pdf_file}: {str(e)}")
        return ""

def download_from_google_drive(url, save_as_filename):
    """
    Download a file from Google Drive and save it locally.
    """
    try:
        # Parse the URL to get the file ID
        parsed_url = urlparse(url)
        
        if 'drive.google.com' in parsed_url.netloc:
            if 'file/d/' in url:
                # Format: https://drive.google.com/file/d/FILE_ID/view
                file_id = url.split('/file/d/')[1].split('/')[0]
            elif 'id=' in url:
                # Format: https://drive.google.com/open?id=FILE_ID
                file_id = parse_qs(parsed_url.query)['id'][0]
            else:
                return {"success": False, "error": "Unsupported Google Drive URL format"}
                
            # Direct download link format
            download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            
            # Download the file
            response = requests.get(download_url, stream=True)
            if response.status_code == 200:
                save_path = os.path.join(RESUMES_DIR, save_as_filename)
                with open(save_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                
                    return {
                        "success": True, 
                    "file_path": save_path
                }
            else:
                return {"success": False, "error": f"Failed to download file: {response.status_code}"}
        else:
            return {"success": False, "error": "Not a Google Drive URL"}
            
    except Exception as e:
        return {"success": False, "error": str(e)}

def save_attachment_as_resume(attachment_path, save_as_filename):
    """
    Copy a LinkedIn attachment to the resumes folder with the desired filename.
    """
    try:
        # Copy the file to the resumes directory with the new name
        destination_path = os.path.join(RESUMES_DIR, save_as_filename)
        shutil.copy2(attachment_path, destination_path)
        return {
            "success": True,
            "file_path": destination_path
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def process_resume(message, resume_filename, sender_name, job_id, analysis):
    """Process a resume from either attachments or Google Drive links."""
    resume_saved = False
    resume_path = None
    resume_info = {}
    
    # Check for LinkedIn attachments first
    if "attachments" in message and not resume_saved:
        for attachment in message["attachments"]:
            file_path = attachment.get("saved_path")
            if file_path and os.path.exists(file_path) and file_path.lower().endswith('.pdf'):
                print(f"Processing PDF attachment: {attachment.get('filename')}")
                
                # Save the attachment to resumes folder with the proper name
                save_result = save_attachment_as_resume(file_path, resume_filename)
                
                if save_result.get("success", False):
                    resume_path = save_result.get("file_path")
                    resume_saved = True
                    
                    # Extract text and analyze resume
                    resume_text = extract_text_from_pdf(resume_path)
                    if resume_text:
                        resume_info = extract_info_from_resume_with_gpt(resume_text, resume_filename)
                    break
    
    # Check for Google Drive links if no attachment was processed
    if "google_drive_links" in message and not resume_saved:
        drive_links = message["google_drive_links"]
        if drive_links:
            drive_link = drive_links[0]  # Process the first link
            print(f"Processing Google Drive link: {drive_link}")
            
            # Download from Google Drive and save with proper name
            drive_result = download_from_google_drive(drive_link, resume_filename)
            
            if drive_result.get("success", False):
                resume_path = drive_result.get("file_path")
                resume_saved = True
                
                # Extract text and analyze resume
                resume_text = extract_text_from_pdf(resume_path)
                if resume_text:
                    resume_info = extract_info_from_resume_with_gpt(resume_text, resume_filename)
    
    # If we successfully processed a resume, prepare referral data
    if resume_saved and resume_info:
        # Prepare the referral data for Excel
        return {
            "job_id": job_id,
            "name": resume_info.get("name", sender_name),
            "email": resume_info.get("email", analysis.get("email", "")),
            "phone": resume_info.get("phone", ""),
            "years_of_experience": resume_info.get("years_of_experience", ""),
            "Resume Path": resume_path,
            "Position": analysis.get("job_position", ""),
            "Company": analysis.get("company", "")
        }
    
    return None

def analyze_resume_with_gpt(resume_text, sender_name, message_content="", filename=""):
    """
    Use Azure OpenAI to analyze resume text and extract key information.
    Returns a dictionary with name, email, phone, years of experience, and job ID.
    """
    try:
        # Print the filename we're analyzing
        print(f"Analyzing resume file: {filename}")
        
        # Truncate if resume is too long
        if len(resume_text) > 10000:
            resume_text = resume_text[:10000] + "..."
        
        # Print debugging information about the message content
        print(f"\nAnalyzing data for: {sender_name}")
        print(f"Message content length: {len(message_content)}")
        
        # Create separate prompts for resume and message
        resume_prompt = f"""
        Extract the following information from this resume:
        
        RESUME TEXT:
        {resume_text}
        
        Please extract:
        1. Full name
        2. Email address
        3. Phone number
        4. Total years of work experience (calculate based on work history)
        
        For calculating years of experience:
        - Sum the total time (in years) across all professional work experiences
        - For current positions, calculate time from start date until present
        - For past positions, calculate time from start date to end date
        - Round to the nearest quarter (0.25, 0.5, 0.75, 1.0)
        - Do NOT include internships, volunteering, or educational experiences unless they were full-time paid professional roles
        - For overlapping time periods, do not double-count the overlapping time
        - If exact dates aren't provided, make a reasonable estimation based on the information available
        - Be consistent: if you calculate 1 year and 2 months, that should be 1.25 years
        
        Return ONLY a JSON object with these keys: name, email, phone, years_of_experience
        Make the response valid JSON with no additional text, no markdown formatting, no code blocks.
        If you can't find a value, use "Not found" as the value.
        """
        
        # Call Azure OpenAI API for resume analysis
        print("Sending resume to Azure OpenAI for analysis...")
        resume_response = client.chat.completions.create(
            model="gpt4o",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that extracts structured information from resumes. Respond with plain JSON only, no markdown formatting or code blocks."},
                {"role": "user", "content": resume_prompt}
            ],
            temperature=0.1
        )
        
        resume_content = resume_response.choices[0].message.content
        
        # Extract job ID from LinkedIn message if available
        job_id = "Not found"
        if message_content:
            # Debug: Print the actual message content for inspection
            print("\nDEBUG - LINKEDIN MESSAGE CONTENT:")
            print("=" * 80)
            print(message_content)
            print("=" * 80)
            
            print("Sending LinkedIn message to Azure OpenAI to extract job ID...")
            
            job_id_prompt = f"""
            Extract ONLY the job ID from this LinkedIn message.
            A job ID is typically mentioned after phrases like "Job ID:", "JobId:", "Position ID:", "Reference #:", etc.
            If multiple job IDs are present, include all of them separated by semicolons.
            
            LINKEDIN MESSAGE:
            {message_content}
            
            Return ONLY a valid JSON with the key "job_id" and the value being the extracted job ID(s).
            If no job ID is found, return {{"job_id": "Not found"}}
            """
            
            job_id_response = client.chat.completions.create(
                model="gpt4o",
                messages=[
                    {"role": "system", "content": "You are a precise assistant that extracts job IDs from messages. Return only valid JSON with a single key 'job_id'."},
                    {"role": "user", "content": job_id_prompt}
                ],
                temperature=0.1,
                max_tokens=500  # Ensure we get a complete response
            )
            
            job_id_content = job_id_response.choices[0].message.content
            print(f"Job ID extraction response: {job_id_content}")
            
            try:
                # Clean the response to ensure it's valid JSON
                cleaned_content = job_id_content.strip()
                if cleaned_content.startswith("```"):
                    cleaned_content = re.sub(r'^```.*\n', '', cleaned_content)
                    cleaned_content = re.sub(r'\n```$', '', cleaned_content)
                
                job_id_data = json.loads(cleaned_content)
                job_id = job_id_data.get("job_id", "Not found")
                print(f"Extracted job ID: {job_id}")
            except Exception as e:
                print(f"Error parsing job ID response: {str(e)}")
                
                # Try a regex backup method if JSON parsing fails
                print("Attempting regex backup for job ID extraction...")
                job_id_matches = re.findall(r'(?:Job\s*ID|JobId):\s*([A-Za-z0-9-]+(?:,\s*[A-Za-z0-9-]+)*)', message_content, re.IGNORECASE)
                if job_id_matches:
                    job_id = job_id_matches[0].replace(',', ';')
                    print(f"Extracted job ID via regex backup: {job_id}")
        else:
            print("No message content available for job ID extraction.")
        
        # Parse the resume JSON response
        try:
            # Clean up any markdown formatting
            cleaned_content = resume_content.strip()
            if cleaned_content.startswith("```"):
                cleaned_content = re.sub(r'^```.*\n', '', cleaned_content)
                cleaned_content = re.sub(r'\n```$', '', cleaned_content)
            
            print(f"Cleaned resume data: {cleaned_content[:100]}...")
            resume_data = json.loads(cleaned_content)
            
            # Combine resume data with job ID
            result = {
                "name": resume_data.get("name", "Not found"),
                "email": resume_data.get("email", "Not found"),
                "phone": resume_data.get("phone", "Not found"),
                "years_of_experience": resume_data.get("years_of_experience", "Not found"),
                "job_id": job_id
            }
            
            # Ensure we have all required fields
            required_fields = ["name", "email", "phone", "years_of_experience", "job_id"]
            for field in required_fields:
                if field not in result or not result[field]:
                    result[field] = "Not found"
            
            print(f"Final analysis result: {result}")
            return result
        except json.JSONDecodeError as e:
            print(f"Error parsing GPT response as JSON: {resume_content}")
            print(f"JSON error: {str(e)}")
            return {
                "name": sender_name,
                "email": "Error parsing API response",
                "phone": "Error parsing API response",
                "years_of_experience": "Error parsing API response", 
                "job_id": job_id if job_id != "Not found" else "Error parsing API response"
            }
            
    except Exception as e:
        print(f"Error analyzing resume: {str(e)}")
        return {
            "name": sender_name,
            "email": f"Error: {str(e)}",
            "phone": f"Error: {str(e)}",
            "years_of_experience": f"Error: {str(e)}",
            "job_id": "Error: {str(e)}"
        }

def process_linkedin_messages():
    """Process LinkedIn messages with referral requests and resumes."""
    try:
        # Find the most recent referral messages file
        messages_file = get_latest_file("linkedin_referrals")
        if not messages_file:
            print("No LinkedIn referral message files found.")
            return False
        
        # Find all resume files in the resumes directory
        resume_files = glob.glob(os.path.join(RESUMES_DIR, "*.pdf"))
        
        if not resume_files:
            print("No resume files found in the resumes directory.")
            return False
        
        print(f"\nFound message files: {[messages_file]}")
        print(f"Found resume files: {resume_files}")
        
        # Load messages from file
        print(f"\nReading messages from: {messages_file}")
        with open(messages_file, 'r', encoding='utf-8') as f:
            messages_data = json.load(f)
        
        print(f"Loaded {len(messages_data)} conversations from messages file")
        
        # Extract message content by sender for matching
        message_content_by_sender = {}
        original_names = {}  # Map display names to actual names when available
        
        for msg in messages_data:
            display_name = msg.get('sender', 'Unknown')
            sender_key = display_name.replace(' ', '_')
            
            # Check if we found an actual name in the message
            actual_name = msg.get('actual_name')
            if actual_name:
                # Use the actual name as the key instead
                actual_key = actual_name.replace(' ', '_')
                print(f"Using actual name '{actual_name}' instead of '{display_name}' for matching")
                sender_key = actual_key
                original_names[sender_key] = display_name  # Store mapping for later
            
            # Combine all message content from this sender
            full_content = ""
            for message in msg.get('messages', []):
                content = message.get('content', '')
                if content:
                    full_content += content + "\n"
            
            if full_content:
                message_content_by_sender[sender_key] = full_content
        
        # Debug first message structure
        if messages_data:
            first_msg = messages_data[0]
            print("\nDEBUG - First message structure:")
            print(f"  sender: {first_msg.get('sender', 'Unknown')}")
            print(f"  preview: {first_msg.get('preview', 'No preview')}")
            print(f"  is_potential_referral: {first_msg.get('is_potential_referral', False)}")
            print(f"  has_resume_downloaded: {first_msg.get('has_resume_downloaded', False)}")
            print(f"  messages: Array with {len(first_msg.get('messages', []))} items")
        
        # Extract and show available message content for debugging
        for sender, content in message_content_by_sender.items():
            print(f"Extracted content for {sender}, length: {len(content)} chars")
            print(f"First 100 chars: {content[:100]}...")
        
        print(f"\nAvailable senders in messages file: {list(message_content_by_sender.keys())}")
        
        # Process only the resume files from referral requests
        analyses = []
        print(f"\nProcessing {len(resume_files)} resumes...")
        
        for i, resume_file in enumerate(resume_files):
            print("\n" + "=" * 60)
            print(f"RESUME ANALYSIS {i+1}/{len(resume_files)}")
            print("=" * 60)
            print(f"File: {os.path.basename(resume_file)}")
            
            # Extract sender name from filename (e.g., John_Smith_resume.pdf -> John_Smith)
            file_basename = os.path.basename(resume_file)
            
            # Handle different possible filename patterns
            if '_resume' in file_basename:
                sender_name = file_basename.split('_resume')[0]
            elif '_cv' in file_basename.lower():
                sender_name = file_basename.split('_cv')[0]
            else:
                # Just remove the extension
                sender_name = os.path.splitext(file_basename)[0]
            
            print(f"Extracted sender name from filename: {sender_name}")
            
            # Find the corresponding message content
            message_content = None
            print("-" * 40)
            print("MESSAGE MATCHING")
            print("-" * 40)
            
            # Try direct match first
            if sender_name in message_content_by_sender:
                message_content = message_content_by_sender[sender_name]
                print(f"✓ Found direct match for message content with key: {sender_name}")
            
            # If no direct match, try partial match
            if not message_content:
                best_match = None
                best_match_score = 0
                
                for key in message_content_by_sender.keys():
                    # Compare by cleaning up both names and checking partial matches
                    clean_sender = sender_name.lower().replace('_', ' ')
                    clean_key = key.lower().replace('_', ' ')
                    
                    if clean_sender in clean_key or clean_key in clean_sender:
                        # Calculate match score based on common parts
                        sender_parts = set(clean_sender.split())
                        key_parts = set(clean_key.split())
                        common_parts = sender_parts.intersection(key_parts)
                        match_score = len(common_parts) / max(len(sender_parts), len(key_parts))
                        
                        if match_score > best_match_score:
                            best_match = key
                            best_match_score = match_score
                
                if best_match:
                    message_content = message_content_by_sender[best_match]
                    print(f"✓ Found partial match for message content with key: {best_match}")
                    print(f"  Match quality: {best_match_score:.2f} (higher is better)")
            
            # Try name parts match if we still don't have a match
            if not message_content:
                print("Trying name parts match...")
                # Convert underscores in sender_name to spaces for comparison
                sender_parts = sender_name.replace("_", " ").lower().split()
                best_match = None
                best_match_count = 0
                
                for key in message_content_by_sender.keys():
                    key_parts = key.replace("_", " ").lower().split()
                    # Count how many parts match
                    matching_parts = sum(1 for part in sender_parts if part in key_parts)
                    if matching_parts > best_match_count:
                        best_match = key
                        best_match_count = matching_parts
                
                if best_match and best_match_count > 0:
                    message_content = message_content_by_sender[best_match]
                    print(f"✓ Found name parts match with key: {best_match}")
                    print(f"  Matching parts: {best_match_count}")
            
            if not message_content:
                print(f"❌ WARNING: No message content found for {sender_name}")
            
            # Extract text from PDF
            print("-" * 40)
            print("PDF TEXT EXTRACTION")
            print("-" * 40)
            resume_text = extract_text_from_pdf(resume_file)
            
            if not resume_text:
                print(f"❌ Could not extract text from {resume_file}")
                continue
            else:
                print(f"✓ Successfully extracted {len(resume_text)} characters from PDF")
                
            # Analyze resume with GPT
            print("-" * 40)
            print("GPT ANALYSIS")
            print("-" * 40)
            print(f"Sending resume to Azure OpenAI for analysis...")
            analysis = analyze_resume_with_gpt(resume_text, sender_name, message_content, os.path.basename(resume_file))
            
            print("-" * 40)
            print("ANALYSIS RESULTS")
            print("-" * 40)
            print(f"Name: {analysis.get('name', 'Not found')}")
            print(f"Email: {analysis.get('email', 'Not found')}")
            print(f"Phone: {analysis.get('phone', 'Not found')}")
            print(f"Experience: {analysis.get('years_of_experience', 'Not found')} years")
            print(f"Job ID: {analysis.get('job_id', 'Not found')}")
            
            # Add to our analyses list
            analyses.append(analysis)
            print(f"✓ Added to Excel report queue")
            
            # Wait a bit to avoid rate limits
            time.sleep(1)
        
        if not analyses:
            print("\n" + "=" * 60)
            print("ERROR: NO ANALYSES COMPLETED")
            print("=" * 60)
            print("No resume analyses were completed.")
            return False
        
        # Create Excel file with analyses
        print("\n" + "=" * 60)
        print("CREATING EXCEL REPORT")
        print("=" * 60)
        create_excel_report(analyses)
        return True
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("ERROR PROCESSING MESSAGES")
        print("=" * 60)
        print(f"Error: {str(e)}")
        return False

def create_excel_report(analyses):
    """Create an Excel report with resume analyses."""
    try:
        # Print received analyses for debugging
        print(f"Creating Excel report with {len(analyses)} analyses")
        print("-" * 40)
        print("DATA FOR EXCEL REPORT")
        print("-" * 40)
        for i, analysis in enumerate(analyses):
            print(f"Entry {i+1}:")
            print(f"  Name: {analysis.get('name', 'Not found')}")
            print(f"  Email: {analysis.get('email', 'Not found')}")
            print(f"  Job ID: {analysis.get('job_id', 'Not found')}")
        
        # Create DataFrame
        print("\nCreating DataFrame...")
        df = pd.DataFrame(analyses)
        
        # Debug column names
        print(f"DataFrame columns: {df.columns.tolist()}")
        
        # Ensure all required columns exist with correct case
        required_columns = ["name", "email", "phone", "years_of_experience", "job_id"]
        for col in required_columns:
            if col not in df.columns:
                print(f"Column {col} not found in data, adding empty column")
                df[col] = "Not found"
        
        # Reorder columns
        df = df[required_columns]
        
        # Create a styled Excel file
        print("\nFormatting Excel workbook...")
        wb = Workbook()
        ws = wb.active
        
        # Add header with proper capitalization for display
        header_display = ["Name", "Email", "Phone", "Years of Experience", "Job ID"]
        ws.append(header_display)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        print("Adding data rows...")
        for analysis in analyses:
            ws.append([
                analysis.get("name", "Not found"),
                analysis.get("email", "Not found"),
                analysis.get("phone", "Not found"),
                analysis.get("years_of_experience", "Not found"),
                analysis.get("job_id", "Not found")
            ])
        
        # Auto-adjust column width
        print("Adjusting column widths...")
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook with the new name
        excel_file = "linkedin_referrals.xlsx"
        print(f"Saving Excel file to {excel_file}...")
        wb.save(excel_file)
        print(f"✅ Saved Excel report to {excel_file}")
        
        print("✅ Excel report created successfully!")
        return True
    except Exception as e:
        print(f"❌ Error creating Excel report: {str(e)}")
        return False

if __name__ == "__main__":
    # Run the process directly
    result = process_linkedin_messages()
    if result:
        print("Processing complete!")
    else:
        print("Processing failed or no referral requests with resumes were found.") 